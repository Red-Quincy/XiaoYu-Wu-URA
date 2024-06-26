#!/usr/bin/env python
# coding: utf-8

# In[8]:


from openpyxl import Workbook, load_workbook #openpyxl does all the excel work
from natsort import natsorted
import os #is for something i found online for sorting files
import xlrd #The BC data only has xls
import xlwt
from xlutils.copy import copy
from datetime import datetime

path=input('Please enter the path traffic control station at a specific year')
saved_Path = str(path)
path2 = str(path)
path_list = path.split() 
path_list.insert(0,'"')
path_list.insert(0,'r')
path_list.insert(len(path_list),'"') #this adds r" and " to change the path. It has to be this way otherwise it doesnt work

information = path_list[-2]
year_list = [str(information[-4]),str(information[-3]),str(information[-2]),str(information[-1])]
counter_list = [str(information[-11]), str(information[-10]), str(information[-9]), str(information[-8]), str(information[-7]),str(information[-6])]

path = ''.join(path_list)
year = ''.join(year_list)
counterStation = ''.join(counter_list)

stations = ['P152EW','P162EW','P167EW','P171EW','P174EW','P179EW','P701EW','P711EW','P1600EW','P1622EW','P1799EW','P91010EW']
respectiveTruckPercent = [0.07,0.011,0.028,0.024,0.068,0.142,0.109,0.107,0.1,0.128,0.168,0.136,0.051] #these need to be updated

truck_percent = respectiveTruckPercent[stations.index(counterStation)]

#STOP POINT 
def excel_round(number, digits):
    factor = 10 ** digits
    rounded_number = round(number * factor)
    return rounded_number / factor
#DEFINING A FUNCTION THAT WILL BE USED LATER

def day_of_week(year, month, day):
    # Create a datetime object for the given date
    date_obj = datetime(year, month, day)
    
    # Get the day of the week as an integer (0 = Monday, 1 = Tuesday, ..., 6 = Sunday)
    day_index = date_obj.weekday()
    
    # Return the day of the week as a string
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    return days[day_index]

# Example usage
#year = 2024
#month = 7
#day = 18
#print(day_of_week(year, month, day))  # Output will be the day of the week for March 5, 2024

#STOP POINT

def compile_workbooks(workbooks_path, final_filename):
    if not isinstance(workbooks_path, str):
        raise TypeError("Argument workbooks_path must be of type str.")
    
    if not isinstance(final_filename, str):
        raise TypeError("Argument final_filename must be of type str.")

    if not os.path.exists(workbooks_path):
        raise NotADirectoryError("Argument workbook_path is not a directory.")

    if not final_filename.endswith(".xlsx"):
        raise ValueError('final_filename must end with the string ".xlsx"')
        
    if final_filename in os.listdir(workbooks_path):
        raise ValueError(f'There is already a file named {final_filename} in {workbooks_path}. '
                         f'Remove this file first or change the final_filename parameter value.')
    
    wbs = []
    
    files = natsorted(os.listdir(workbooks_path))
    for file in files:
        if not file.startswith("~$") and file.endswith(".xls"):
            wb = xlrd.open_workbook(os.path.join(workbooks_path, file))
            wbs.append(wb)
    
    final_wb = Workbook()
    
    titles = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
    
    # Set sheet titles and copy data from the first sheet of each workbook
    for index, wb in enumerate(wbs):
        if index >= len(titles):
            break
        ws1 = wb.sheet_by_index(0)
        final_ws = final_wb.create_sheet(title=titles[index])
        
        # Copy data
        for i in range(ws1.nrows):
            for j in range(ws1.ncols):
                final_ws.cell(row=i+1, column=j+1, value=ws1.cell_value(i, j))

    # Remove the default "Sheet" sheet
    default_sheet = final_wb["Sheet"]
    final_wb.remove(default_sheet)

    final_wb.save(os.path.join(workbooks_path, final_filename))

#workbooks_path = r'C:\Users\redqu\OneDrive\Waterloo\2AB\XiaoYu URA\Alberta\60021540\2021'

path_split = saved_Path.split()
final = "\\"+ str(year) + 'FullYearResults.xlsx'
path_split.append(final)
final_path = ''.join(path_split) #Setting where the file should be saved

FP = final_path.split() 
FP.insert(0,'"')
FP.insert(0,'r')
FP.insert(len(path_list),'"')
#final_path = ''.join(FP)

if __name__ == '__main__':
    compile_workbooks(path2, final_path)

#STOP POINT

titles = months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
workbook = load_workbook(final_path) #USE THIS CODE EXACTLY LIKE THIS TO LOAD THE RIGHT WORKBOOK
rowMax = [0]*12

for month in titles: 
    workbook.active = workbook[month]
    sheet = workbook.active
    index = titles.index(month)
    k=12
    while True: 
        if sheet.cell(row=k,column=2).value != None: 
            k+=1
            continue
        else: 
            k-=1
            break
    rowMax[index]=k


workbook = load_workbook(final_path)
workbook.active = workbook['jan']
sheet = workbook.active

days_of_the_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
hours_of_the_week_data_points =[0]*24
days_of_the_week_data_points = [0]*7
monday = [0] * 24
tuesday = [0] * 24
wednesday = [0] * 24
thursday = [0] * 24
friday = [0] * 24
saturday = [0] * 24
sunday = [0] * 24
days_of_the_week_values = [monday,tuesday,wednesday,thursday,friday,saturday,sunday]

for month in titles: 
    workbook.active = workbook[month]
    titlesIndex = titles.index(month)
    finalRow = rowMax[titlesIndex]
    sheet = workbook.active
    row_thing=12
    
    d=1

    for i in range(12,rowMax[titlesIndex]+1):
        day = day_of_week(int(year), titlesIndex+1, d)
        index = days_of_the_week.index(day)
        days_of_the_week_data_points[index] += 1
        hour = 0
        for k in range(3,27):
            if sheet.cell(row=i,column=k).value is not None:
                hours_of_the_week_data_points[hour] += 1
                days_of_the_week_values[index][hour] += int((sheet.cell(row=i,column=k).value)*truck_percent)
                hour+=1
        d+=1
        

days_of_the_week_average = []

w = 0
for sublist in days_of_the_week_values:
    average_sublist = []  
    hour = 0
    for data_point in sublist:
        daily_average = int(data_point / hours_of_the_week_data_points[hour])  # Calculate the daily average for each data point
        hour += 1
        average_sublist.append(daily_average)  # Add the daily average to the sublist
    days_of_the_week_average.append(average_sublist) 
    w += 1 

    
workbook = load_workbook(final_path)
dataPoints = (rowMax[6]-11)+(rowMax[7]-11)
SADTT_volumeCount = 0
SADT_volumeCount = 0
error = False


for i in range(6,8): 
    workbook.active = workbook[titles[i]]
    finalRow = rowMax[i]
    sheet = workbook.active
    for j in range(12,rowMax[i]+1):
        error = False
        day = day_of_week(int(year),i+1,j-11)
        for k in range(3,27):
            if sheet.cell(row=j,column=k).value is None:
                error = True 
                break
        if error is False: 
            for k in range(3,27):
                SADTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                SADT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
        else: 
            dataPoints -= 1
            
SADTT_volumeCount = excel_round(SADTT_volumeCount/dataPoints,0)
SADT_volumeCount = excel_round(SADT_volumeCount/dataPoints,0)

#SADT works yep man yepppers

workbook = load_workbook(final_path)
dataPoints = 0
SAWDTT_volumeCount = 0
SAWDT_volumeCount = 0

for i in range(6,8): 
    workbook.active = workbook[titles[i]]
    finalRow = rowMax[i]
    sheet = workbook.active 
    
    for j in range(12,rowMax[i]+1):
        error = False 
        day = day_of_week(int(year),i+1,j-11)
        if day != 'Saturday' and day != 'Sunday':
            dataPoints += 1
            error = False
            for k in range(3,27):
                if sheet.cell(row=j,column=k).value is None: 
                    error = True 
            if error is False: 
                for k in range(3,27):
                    SAWDTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                    SAWDT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
            else: 
                dataPoints -= 1
        else: 
            continue            

SAWDTT_volumeCount = excel_round(SAWDTT_volumeCount/dataPoints,0)
SAWDT_volumeCount = excel_round(SAWDT_volumeCount/dataPoints,0)

#SAWDT does in fact work thankfully.




#WADT 
    #Winter average daily traffic. The average twenty-four-hour, two-way traffic 
    #for the period January 1st to March 31st, plus December 1st to December 31st, including weekends.   

workbook = load_workbook(final_path)
WADTT_volumeCount = 0
WADT_volumeCount = 0 

dataPoints = 0

monthsIndex = [0,1,2,11]



for i in monthsIndex: 
    workbook.active = workbook[titles[i]]
    finalRow = rowMax[i]
    sheet = workbook.active

    for j in range(12,rowMax[i]+1):
        error = False
        dataPoints += 1
        for k in range(3,27):
            if sheet.cell(row=j,column=k).value is None:
                #dataPoints -= 1 
                error = True 
                break
        if error is False: 
            for k in range(3,27):
                WADTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                WADT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
        else: 
            dataPoints -= 1

WADTT_volumeCount = excel_round(WADTT_volumeCount/dataPoints,0)
WADT_volumeCount = excel_round(WADT_volumeCount/dataPoints,0)

#THIS WILL ALSO WORK YEPPERS.

#AADTT  

workbook = load_workbook(final_path)
dataPoints = 0
AADTT_volumeCount = 0
AADT_volumeCount = 0 

for i in range(12): 
    workbook.active = workbook[titles[i]]
    finalRow = rowMax[i]
    sheet = workbook.active

    for j in range(12,rowMax[i]+1):
        dataPoints += 1
        error = False 
        for k in range(3,27):
            if sheet.cell(row=j,column=k).value is None:
                error = True 
                break
        if error is False: 
            for k in range(3,27):
                AADTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                AADT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
        else: 
            dataPoints -= 1


AADTT_volumeCount = excel_round(AADTT_volumeCount/dataPoints,0)
AADT_volumeCount = excel_round(AADT_volumeCount/dataPoints,0)

#AADT does work yep man yeppers


#AT LAST
wb = Workbook()
ws = wb.active

ws.title = "Results"
ws['A1'] = "Traffic Control Station"
ws['B1'] = counterStation
ws['B6'] = "AADT(T)"
ws['C6'] = "SADT(T)"
ws['D6'] = "SAWDT(T)"
ws['e6'] = "WADT(T)"
ws['a7'] = "Truck"
ws['a8'] = "Regular"
ws['a9'] = "Truck as % of regular"

ws['b7'] = AADTT_volumeCount
ws['c7'] = SADTT_volumeCount
ws['d7'] = SAWDTT_volumeCount
ws['e7'] = WADTT_volumeCount

ws['B8'] = AADT_volumeCount
ws['c8'] = SADT_volumeCount
ws['d8'] = SAWDT_volumeCount
ws['e8'] = WADT_volumeCount

ws['b9'] = ws['b7'].value/ws['b8'].value
ws['c9'] = ws['c7'].value/ws['c8'].value
ws['d9'] = ws['d7'].value/ws['d8'].value
ws['e9'] = ws['e7'].value/ws['e8'].value

# Specify the path where you want to save the file
placeholder = r"CHANGE THIS TO THE PATH TO WHERE YOU WANT THE RESULTS TO GO\ 60021540-2022Results.xlsx"
placeholder = placeholder.split()
thing = counterStation,'-',year,'Results.xlsx'
placeholder[-1]=''.join(thing)

file_path = ''.join(placeholder)

filePath = r'{}'.format(file_path)


wb.save(filename=r'{}'.format(file_path))


# In[12]:


#THIS IS THE SAME CODE, EXCEPT IT WONT CREATE THE NEW EXCEL FILE LIKE THE OLD CODE DOES. NICE.


from openpyxl import Workbook, load_workbook #openpyxl does all the excel work
from natsort import natsorted
import os #is for something i found online for sorting files
import xlrd #The BC data only has xls
import xlwt
from xlutils.copy import copy
from datetime import datetime

path=input('Please enter the path traffic control station at a specific year')
saved_Path = str(path)
path2 = str(path)
path_list = path.split() 
path_list.insert(0,'"')
path_list.insert(0,'r')
path_list.insert(len(path_list),'"') #this adds r" and " to change the path. It has to be this way otherwise it doesnt work

information = path_list[-2]
year_list = [str(information[-4]),str(information[-3]),str(information[-2]),str(information[-1])]
counter_list = [str(information[-11]), str(information[-10]), str(information[-9]), str(information[-8]), str(information[-7]),str(information[-6])]

path = ''.join(path_list)
year = ''.join(year_list)
counterStation = ''.join(counter_list)

stations = ['P152EW','P162EW','P167EW','P171EW','P174EW','P179EW','P701EW','P711EW','P1600EW','P1622EW','P1799EW','P91010EW']
respectiveTruckPercent = [0.07,0.011,0.028,0.024,0.068,0.142,0.109,0.107,0.1,0.128,0.168,0.136,0.051] #these need to be updated

truck_percent = respectiveTruckPercent[stations.index(counterStation)]

#STOP POINT 
def excel_round(number, digits):
    factor = 10 ** digits
    rounded_number = round(number * factor)
    return rounded_number / factor
#DEFINING A FUNCTION THAT WILL BE USED LATER

def day_of_week(year, month, day):
    # Create a datetime object for the given date
    date_obj = datetime(year, month, day)
    
    # Get the day of the week as an integer (0 = Monday, 1 = Tuesday, ..., 6 = Sunday)
    day_index = date_obj.weekday()
    
    # Return the day of the week as a string
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    return days[day_index]

# Example usage
#year = 2024
#month = 7
#day = 18
#print(day_of_week(year, month, day))  # Output will be the day of the week for March 5, 2024

#STOP POINT
'''
def compile_workbooks(workbooks_path, final_filename):
    if not isinstance(workbooks_path, str):
        raise TypeError("Argument workbooks_path must be of type str.")
    
    if not isinstance(final_filename, str):
        raise TypeError("Argument final_filename must be of type str.")

    if not os.path.exists(workbooks_path):
        raise NotADirectoryError("Argument workbook_path is not a directory.")

    if not final_filename.endswith(".xlsx"):
        raise ValueError('final_filename must end with the string ".xlsx"')
        
    if final_filename in os.listdir(workbooks_path):
        raise ValueError(f'There is already a file named {final_filename} in {workbooks_path}. '
                         f'Remove this file first or change the final_filename parameter value.')
    
    wbs = []
    
    files = natsorted(os.listdir(workbooks_path))
    for file in files:
        if not file.startswith("~$") and file.endswith(".xls"):
            wb = xlrd.open_workbook(os.path.join(workbooks_path, file))
            wbs.append(wb)
    
    final_wb = Workbook()
    
    titles = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
    
    # Set sheet titles and copy data from the first sheet of each workbook
    for index, wb in enumerate(wbs):
        if index >= len(titles):
            break
        ws1 = wb.sheet_by_index(0)
        final_ws = final_wb.create_sheet(title=titles[index])
        
        # Copy data
        for i in range(ws1.nrows):
            for j in range(ws1.ncols):
                final_ws.cell(row=i+1, column=j+1, value=ws1.cell_value(i, j))

    # Remove the default "Sheet" sheet
    default_sheet = final_wb["Sheet"]
    final_wb.remove(default_sheet)

    final_wb.save(os.path.join(workbooks_path, final_filename))
'''
#workbooks_path = r'C:\Users\redqu\OneDrive\Waterloo\2AB\XiaoYu URA\Alberta\60021540\2021'

path_split = saved_Path.split()
final = "\\"+ str(year) + 'FullYearResults.xlsx'
path_split.append(final)
final_path = ''.join(path_split) #Setting where the file should be saved

FP = final_path.split() 
FP.insert(0,'"')
FP.insert(0,'r')
FP.insert(len(path_list),'"')
#final_path = ''.join(FP)

#if __name__ == '__main__':
#    compile_workbooks(path2, final_path)

#STOP POINT

titles = months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
workbook = load_workbook(final_path) #USE THIS CODE EXACTLY LIKE THIS TO LOAD THE RIGHT WORKBOOK
rowMax = [0]*12

for month in titles: 
    workbook.active = workbook[month]
    sheet = workbook.active
    index = titles.index(month)
    k=12
    while True: 
        if sheet.cell(row=k,column=2).value != None: 
            k+=1
            continue
        else: 
            k-=1
            break
    rowMax[index]=k

#print(rowMax)

workbook = load_workbook(final_path)
workbook.active = workbook['jan']
sheet = workbook.active

days_of_the_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
hours_of_the_week_data_points =[0]*24
days_of_the_week_data_points = [0]*7
monday = [0] * 24
tuesday = [0] * 24
wednesday = [0] * 24
thursday = [0] * 24
friday = [0] * 24
saturday = [0] * 24
sunday = [0] * 24
days_of_the_week_values = [monday,tuesday,wednesday,thursday,friday,saturday,sunday]

for month in titles: 
    workbook.active = workbook[month]
    titlesIndex = titles.index(month)
    finalRow = rowMax[titlesIndex]
    sheet = workbook.active
    row_thing=12
    
    d=1

    for i in range(12,rowMax[titlesIndex]+1):
        day = day_of_week(int(year), titlesIndex+1, d)
        index = days_of_the_week.index(day)
        days_of_the_week_data_points[index] += 1
        hour = 0
        for k in range(3,27):
            if sheet.cell(row=i,column=k).value is not None:
                hours_of_the_week_data_points[hour] += 1
                days_of_the_week_values[index][hour] += int((sheet.cell(row=i,column=k).value)*truck_percent)
                hour+=1
        d+=1
        

days_of_the_week_average = []

w = 0
for sublist in days_of_the_week_values:
    average_sublist = []  
    hour = 0
    for data_point in sublist:
        daily_average = int(data_point / hours_of_the_week_data_points[hour])  # Calculate the daily average for each data point
        hour += 1
        average_sublist.append(daily_average)  # Add the daily average to the sublist
    days_of_the_week_average.append(average_sublist) 
    w += 1 

    
workbook = load_workbook(final_path)
dataPoints = (rowMax[6]-11)+(rowMax[7]-11)
SADTT_volumeCount = 0
SADT_volumeCount = 0
error = False


for i in range(6,8): 
    workbook.active = workbook[titles[i]]
    finalRow = rowMax[i]
    sheet = workbook.active
    for j in range(12,rowMax[i]+1):
        error = False
        day = day_of_week(int(year),i+1,j-11)
        for k in range(3,27):
            if sheet.cell(row=j,column=k).value is None:
                error = True 
                break
        if error is False: 
            for k in range(3,27):
                SADTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                SADT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
        else: 
            dataPoints -= 1
            
SADTT_volumeCount = excel_round(SADTT_volumeCount/dataPoints,0)
SADT_volumeCount = excel_round(SADT_volumeCount/dataPoints,0)

#SADT works yep man yepppers

workbook = load_workbook(final_path)
dataPoints = 0
SAWDTT_volumeCount = 0
SAWDT_volumeCount = 0

for i in range(6,8): 
    workbook.active = workbook[titles[i]]
    finalRow = rowMax[i]
    sheet = workbook.active 
    
    for j in range(12,rowMax[i]+1):
        error = False 
        day = day_of_week(int(year),i+1,j-11)
        if day != 'Saturday' and day != 'Sunday':
            dataPoints += 1
            error = False
            for k in range(3,27):
                if sheet.cell(row=j,column=k).value is None: 
                    error = True 
            if error is False: 
                for k in range(3,27):
                    SAWDTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                    SAWDT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
            else: 
                dataPoints -= 1
        else: 
            continue            

SAWDTT_volumeCount = excel_round(SAWDTT_volumeCount/dataPoints,0)
SAWDT_volumeCount = excel_round(SAWDT_volumeCount/dataPoints,0)

#SAWDT does in fact work thankfully.




#WADT 
    #Winter average daily traffic. The average twenty-four-hour, two-way traffic 
    #for the period January 1st to March 31st, plus December 1st to December 31st, including weekends.   

workbook = load_workbook(final_path)
WADTT_volumeCount = 0
WADT_volumeCount = 0 

dataPoints = 0

monthsIndex = [0,1,2,11]



for i in monthsIndex: 
    workbook.active = workbook[titles[i]]
    finalRow = rowMax[i]
    sheet = workbook.active

    for j in range(12,rowMax[i]+1):
        error = False
        dataPoints += 1
        for k in range(3,27):
            if sheet.cell(row=j,column=k).value is None:
                #dataPoints -= 1 
                error = True 
                break
        if error is False: 
            for k in range(3,27):
                WADTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                WADT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
        else: 
            dataPoints -= 1

WADTT_volumeCount = excel_round(WADTT_volumeCount/dataPoints,0)
WADT_volumeCount = excel_round(WADT_volumeCount/dataPoints,0)

#THIS WILL ALSO WORK YEPPERS.

#AADTT  

workbook = load_workbook(final_path)
dataPoints = 0
AADTT_volumeCount = 0
AADT_volumeCount = 0 

for i in range(12): 
    workbook.active = workbook[titles[i]]
    finalRow = rowMax[i]
    sheet = workbook.active

    for j in range(12,rowMax[i]+1):
        dataPoints += 1
        error = False 
        for k in range(3,27):
            if sheet.cell(row=j,column=k).value is None:
                error = True 
                break
        if error is False: 
            for k in range(3,27):
                AADTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                AADT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
        else: 
            dataPoints -= 1


AADTT_volumeCount = excel_round(AADTT_volumeCount/dataPoints,0)
AADT_volumeCount = excel_round(AADT_volumeCount/dataPoints,0)

#AADT does work yep man yeppers


#AT LAST
wb = Workbook()
ws = wb.active

ws.title = "Results"
ws['A1'] = "Traffic Control Station"
ws['B1'] = counterStation
ws['B6'] = "AADT(T)"
ws['C6'] = "SADT(T)"
ws['D6'] = "SAWDT(T)"
ws['e6'] = "WADT(T)"
ws['a7'] = "Truck"
ws['a8'] = "Regular"
ws['a9'] = "Truck as % of regular"

ws['b7'] = AADTT_volumeCount
ws['c7'] = SADTT_volumeCount
ws['d7'] = SAWDTT_volumeCount
ws['e7'] = WADTT_volumeCount

ws['B8'] = AADT_volumeCount
ws['c8'] = SADT_volumeCount
ws['d8'] = SAWDT_volumeCount
ws['e8'] = WADT_volumeCount

ws['b9'] = ws['b7'].value/ws['b8'].value
ws['c9'] = ws['c7'].value/ws['c8'].value
ws['d9'] = ws['d7'].value/ws['d8'].value
ws['e9'] = ws['e7'].value/ws['e8'].value

# Specify the path where you want to save the file
placeholder = r"CHANGE THIS TO THE PATH TO WHERE YOU WANT THE RESULTS TO GO\ 60021540-2022Results.xlsx"
placeholder = placeholder.split()
thing = counterStation,'-',year,'Results.xlsx'
placeholder[-1]=''.join(thing)

file_path = ''.join(placeholder)

filePath = r'{}'.format(file_path)


wb.save(filename=r'{}'.format(file_path))


# In[ ]:




