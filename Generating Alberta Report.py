#!/usr/bin/env python
# coding: utf-8

# In[29]:


from openpyxl import Workbook, load_workbook #openpyxl does all the excel work
from natsort import natsorted
import copy
import os #is for something i found online for sorting files

path=input('Please enter the path traffic control station at a specific year')
saved_Path = str(path)
path2 = str(path)
path_list = path.split() 
path_list.insert(0,'"')
path_list.insert(0,'r')
path_list.insert(len(path_list),'"') #this adds r" and " to change the path. It has to be this way otherwise it doesnt work

information = path_list[-2]
year_list = [str(information[-4]),str(information[-3]),str(information[-2]),str(information[-1])] #extracting the year
counter_list = [str(information[-13]),str(information[-12]),str(information[-11]),str(information[-10]),str(information[-9]),str(information[-8]),str(information[-7]),str(information[-6])] #extracting the counter station

path = ''.join(path_list) #gathering variable with the path in the form r"XXXX"
year = ''.join(year_list) #year
counterStation = ''.join(counter_list) #the number of the counter station

stations = ['60021540','60021530','60021520','60021510','60021810','60022250','60022410','60022460','60022610','60022660','60022850','60023010','60023210']
respectiveTruckPercent = [0.043,0.011,0.028,0.024,0.068,0.142,0.109,0.107,0.1,0.128,0.168,0.136,0.051]

truck_percent = respectiveTruckPercent[stations.index(counterStation)]

#STOP POINT 
def excel_round(number, digits):
    factor = 10 ** digits
    rounded_number = round(number * factor)
    return rounded_number / factor
#DEFINING A FUNCTION THAT WILL BE USED LATER
#STOP POINT

def compile_workbooks(workbooks_path, final_filename):
    if not isinstance(workbooks_path, str):
        raise TypeError("Argument workbooks_path must be of type str.")
    
    if not isinstance(final_filename, str):
        raise TypeError("Argument final_filename must be of type str.")

    if not os.path.exists(workbooks_path):
        raise NotADirectoryError("Argument workbook_path is not a directory.")

    if not final_filename.endswith(".xlsx"):
        raise ValueError('final_filename must end with the string ".xlsx"')
        
    if final_filename in os.listdir(workbooks_path):
        raise ValueError(f'There is already a file named {final_filename} in {workbooks_path}. '
                         f'Remove this file first or change the final_filename parameter value.')
    
    wbs = []
    
    files = natsorted(os.listdir(workbooks_path))
    for file in files:
        #print(file)
        if not file.startswith("~$") and file.endswith(".xlsx"):
            wb = load_workbook(os.path.join(workbooks_path, file))
            wbs.append(wb)
    
    final_wb = Workbook()
    final_ws = final_wb.worksheets[0]
    
    titles = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
    final_ws.title = titles[0]
    ws1 = final_wb.create_sheet(titles[1])
    ws2 = final_wb.create_sheet(titles[2])
    ws3 = final_wb.create_sheet(titles[3])
    ws4 = final_wb.create_sheet(titles[4])
    ws5 = final_wb.create_sheet(titles[5])
    ws6 = final_wb.create_sheet(titles[6])
    ws7 = final_wb.create_sheet(titles[7])
    ws8 = final_wb.create_sheet(titles[8])
    ws9 = final_wb.create_sheet(titles[9])
    ws10 = final_wb.create_sheet(titles[10])
    ws11 = final_wb.create_sheet(titles[11])

    k=0 
    #print(wbs)
    for wb in wbs:
        final_ws = final_wb.worksheets[k]
        wb1 = wb
        ws1 = wb1.worksheets[0]
        k+=1
        
        for j in range(1,ws1.max_column+1):
            final_ws.cell(row=1, column=j).value = ws1.cell(row=1, column=j).value
        
        current_row = 2

        ws = wb.worksheets[0]
        mr = ws.max_row 
        mc = ws.max_column 

        for i in range (2, mr + 1): 
            for j in range (1, mc + 1): 
                current_cell = ws.cell(row = i, column = j) 
                final_ws.cell(row = current_row, column = j).value = current_cell.value
                
            current_row += 1


    final_wb.save(os.path.join(workbooks_path, final_filename))

#workbooks_path = r'C:\Users\redqu\OneDrive\Waterloo\2AB\XiaoYu URA\Alberta\60021540\2021'

path_split = saved_Path.split()
final = "\\"+ str(year) + 'FullYearResults.xlsx'
path_split.append(final)
final_path = ''.join(path_split) #Setting where the file should be saved

FP = final_path.split() 
FP.insert(0,'"')
FP.insert(0,'r')
FP.insert(len(path_list),'"')
#final_path = ''.join(FP)

if __name__ == '__main__':
    compile_workbooks(path2, final_path)

#STOP POINT

titles = months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
workbook = load_workbook(final_path) #USE THIS CODE EXACTLY LIKE THIS TO LOAD THE RIGHT WORKBOOK
rowMax = [0]*12

for month in titles: 
    workbook.active = workbook[month]
    sheet = workbook.active
    index = titles.index(month)
    k=16
    while True: 
        if sheet.cell(row=k,column=2).value != None: 
            k+=1
            continue
        else: 
            k-=1
            break
    rowMax[index]=k

#print(rowMax)

workbook = load_workbook(final_path)
workbook.active = workbook['jan']
sheet = workbook.active

days_of_the_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
hours_of_the_week_data_points =[0]*24
days_of_the_week_data_points = [0]*7
monday = [0] * 24
tuesday = [0] * 24
wednesday = [0] * 24
thursday = [0] * 24
friday = [0] * 24
saturday = [0] * 24
sunday = [0] * 24
days_of_the_week_values = [monday,tuesday,wednesday,thursday,friday,saturday,sunday]


for month in titles: 
    workbook.active = workbook[month]
    titlesIndex = titles.index(month)
    finalRow = rowMax[titlesIndex]
    sheet = workbook.active
    for value in workbook.active.iter_cols(min_row=17,max_row=finalRow,min_col=4,max_col=4,values_only=True):
        continue 
    row_thing=17
    #print(month)
    
    for day in value: 
        index = days_of_the_week.index(day)
        days_of_the_week_data_points[index] += 1
        hour = 0
        for k in range(6,30):
            if sheet.cell(row=row_thing,column=k).value is not None:
                hours_of_the_week_data_points[hour] += 1
                days_of_the_week_values[index][hour] += int((sheet.cell(row=row_thing,column=k).value)*truck_percent)
                hour+=1
        row_thing+=1


days_of_the_week_average = []
# Iterate through each sublist in the original data
w = 0
for sublist in days_of_the_week_values:
    average_sublist = []  # Initialize an empty sublist for the daily averages
    # Iterate through each data point in the sublist
    hour = 0
    for data_point in sublist:
        daily_average = int(data_point / hours_of_the_week_data_points[hour])  # Calculate the daily average for each data point
        hour += 1
        average_sublist.append(daily_average)  # Add the daily average to the sublist
    days_of_the_week_average.append(average_sublist) 
    w += 1 
    
#STOP POINT



#SADT 
    #Summer average daily traffic; the average twenty-four-hour, two-way traffic 
    #for the period July 1st to August 31st including weekends.
 

workbook = load_workbook(final_path)
dataPoints = (rowMax[6]-16)+(rowMax[7]-16)
SADTT_volumeCount = 0
SADT_volumeCount = 0
row_thing=17
error = False
#cell = 
#print(sheet.cell(row=17,column=13).value*truck_percent)

for i in range(6,8): 
    workbook.active = workbook[titles[i]]
    #index = i
    finalRow = rowMax[i]
    sheet = workbook.active
    #print(excel_round((sheet.cell(row=17,column=13).value*truck_percent),0))
    for j in range(17,rowMax[i]+1):
        #print(SADT_volumeCount)
        error = False
        for k in range(6,30):
            if sheet.cell(row=j,column=k).value is None:
                #dataPoints -= 1 
                error = True 
                break
        if error is False: 
            for k in range(6,30):
            #print(int(excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)))
                SADTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                SADT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
        else: 
            dataPoints -= 1
            
SADTT_volumeCount = excel_round(SADTT_volumeCount/dataPoints,0)
SADT_volumeCount = excel_round(SADT_volumeCount/dataPoints,0)
    

#EXCEL DOES NOT SUM VALUES PROPERLY> EVEN IF YOU CHANGE IT TO HAVE NO DECIMALS IT WILL STILL SUM 
#AS IF THE DECIMAL POINTS WERE STILL THERE

#STOP POINT
#SAWDT 
    #Summer average weekday traffic. Defined as the average, 
    #twenty-four-hour, two-way traffic for the period July 1st to August 31st excluding weekends.   

workbook = load_workbook(final_path)
dataPoints = 0
SAWDTT_volumeCount = 0
SAWDT_volumeCount = 0
row_thing=17

#cell = 
#print(sheet.cell(row=17,column=13).value*truck_percent)

for i in range(6,8): 
    workbook.active = workbook[titles[i]]
    for value in workbook.active.iter_cols(min_row=17,max_row=rowMax[i],min_col=4,max_col=4,values_only=True):
        continue 
    finalRow = rowMax[i]
    sheet = workbook.active
    #print(excel_round((sheet.cell(row=17,column=13).value*truck_percent),0))
    d = 0 
    for day in value:
        if day != 'Saturday' and day != 'Sunday':
            dataPoints += 1
            error = False
            for k in range(6,30):
                if sheet.cell(row=d+17,column=k).value is None: 
                    #dataPoints -= 1 THIS CANT GO HERE
                    error = True 
            if error is False: 
                for k in range(6,30):
                #print(int(excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)))
                    SAWDTT_volumeCount += excel_round(sheet.cell(row=d+17,column=k).value*truck_percent,0)
                    SAWDT_volumeCount += excel_round(sheet.cell(row=d+17,column=k).value,0)
            else: 
                dataPoints -= 1
            d += 1 #THIS IS WHAT I JUST CHANGED RASHSAHSHASHAHAHSHS
        else: 
            d += 1            

SAWDTT_volumeCount = excel_round(SAWDTT_volumeCount/dataPoints,0)
SAWDT_volumeCount = excel_round(SAWDT_volumeCount/dataPoints,0)
#print(dataPoints)
#print(SAWDT_volumeCount)

#EXCEL DOES NOT SUM VALUES PROPERLY> EVEN IF YOU CHANGE IT TO HAVE NO DECIMALS IT WILL STILL SUM 
#AS IF THE DECIMAL POINTS WERE STILL THERE

#STOP POINT

#WADT 
    #Winter average daily traffic. The average twenty-four-hour, two-way traffic 
    #for the period January 1st to March 31st, plus December 1st to December 31st, including weekends.   

workbook = load_workbook(final_path)
#dataPointsInitial = 121 #NEEDS TO BE UPDATED RAHJGSDAJGHSDAJHGDJA
WADTT_volumeCount = 0
WADT_volumeCount = 0 
row_thing=17
dataPoints = 0

monthsIndex = [0,1,2,11]

#cell = 
#print(sheet.cell(row=17,column=13).value*truck_percent)

for i in monthsIndex: 
    workbook.active = workbook[titles[i]]
    finalRow = rowMax[i]
    sheet = workbook.active
    #print(excel_round((sheet.cell(row=17,column=13).value*truck_percent),0))
    for j in range(17,rowMax[i]+1):
        #print(SADT_volumeCount)
        error = False
        dataPoints += 1
        for k in range(6,30):
            if sheet.cell(row=j,column=k).value is None:
                #dataPoints -= 1 
                error = True 
                break
        if error is False: 
            for k in range(6,30):
                #print(int(excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)))
                WADTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                WADT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
        else: 
            dataPoints -= 1

WADTT_volumeCount = excel_round(WADTT_volumeCount/dataPoints,0)
WADT_volumeCount = excel_round(WADT_volumeCount/dataPoints,0)
    
#print(dataPoints)
#print(WADTT_volumeCount)

#EXCEL DOES NOT SUM VALUES PROPERLY> EVEN IF YOU CHANGE IT TO HAVE NO DECIMALS IT WILL STILL SUM 
#AS IF THE DECIMAL POINTS WERE STILL THERE

#STOP POINT


#AADTT  

workbook = load_workbook(final_path)
dataPoints = 0
AADTT_volumeCount = 0
AADT_volumeCount = 0 
row_thing=17

for i in range(12): 
    workbook.active = workbook[titles[i]]
    #index = i
    finalRow = rowMax[i]
    sheet = workbook.active
    #print(excel_round((sheet.cell(row=17,column=13).value*truck_percent),0))
    for j in range(17,rowMax[i]+1):
        dataPoints += 1
        error = False #THIS IS WHAT I CHANGED RAHHHAHSHAHHSHS
        #print(SADT_volumeCount)
        for k in range(6,30):
            if sheet.cell(row=j,column=k).value is None:
                #dataPoints -= 1 
                error = True 
                break
        if error is False: 
            for k in range(6,30):
                #print(int(excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)))
                AADTT_volumeCount += excel_round(sheet.cell(row=j,column=k).value*truck_percent,0)
                AADT_volumeCount += excel_round(sheet.cell(row=j,column=k).value,0)
        else: 
            dataPoints -= 1


AADTT_volumeCount = excel_round(AADTT_volumeCount/dataPoints,0)
AADT_volumeCount = excel_round(AADT_volumeCount/dataPoints,0)
    
#print(dataPoints)
#print(AADTT_volumeCount)

#EXCEL DOES NOT SUM VALUES PROPERLY> EVEN IF YOU CHANGE IT TO HAVE NO DECIMALS IT WILL STILL SUM 
#AS IF THE DECIMAL POINTS WERE STILL THERE

#STOP POINT




#AT LAST
wb = Workbook()

ws = wb.active

ws.title = "Results"
ws['A1'] = "Traffic Control Station"
ws['B1'] = counterStation
ws['B6'] = "AADT(T)"
ws['C6'] = "SADT(T)"
ws['D6'] = "SAWDT(T)"
ws['e6'] = "WADT(T)"
ws['a7'] = "Truck"
ws['a8'] = "Regular"
ws['a9'] = "Truck as % of regular"

ws['b7'] = AADTT_volumeCount
ws['c7'] = SADTT_volumeCount
ws['d7'] = SAWDTT_volumeCount
ws['e7'] = WADTT_volumeCount

ws['B8'] = AADT_volumeCount
ws['c8'] = SADT_volumeCount
ws['d8'] = SAWDT_volumeCount
ws['e8'] = WADT_volumeCount

ws['b9'] = ws['b7'].value/ws['b8'].value
ws['c9'] = ws['c7'].value/ws['c8'].value
ws['d9'] = ws['d7'].value/ws['d8'].value
ws['e9'] = ws['e7'].value/ws['e8'].value

# Specify the path where you want to save the file
placeholder = r"CHANGE THIS TO THE PATH TO WHERE YOU WANT THE RESULTS TO GO\ 60021540-2022Results.xlsx"
placeholder = placeholder.split()
thing = counterStation,'-',year,'Results.xlsx'
placeholder[-1]=''.join(thing)

#print(placeholder)
#spaceProblem = ' '.join(placeholder[0:2])
#del placeholder[1]
#placeholder[0] = spaceProblem
file_path = ''.join(placeholder)

filePath = r'{}'.format(file_path)
#print(filePath)

wb.save(filename=r'{}'.format(file_path))


# In[ ]:




